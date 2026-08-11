#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Dựng workbook so sánh giá vé từ checkpoint JSON.

    python3 build_workbook.py checkpoint.json output.xlsx

Số cột giá sinh động theo meta.nguon (cũng là thứ tự ưu tiên của Giá chốt).
Cột 'Chênh lệch max %' chỉ xuất hiện khi có >= 2 nguồn.
Sau khi chạy: recalc.py, rồi đối chiếu độc lập — xem SKILL.md §7.
"""
import json, sys, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter as L

NA, MONEY, AR = "Không khả dụng", '#,##0 "₫"', 'Arial'
F_TITLE = Font(name=AR, size=14, bold=True)
F_GRP   = Font(name=AR, size=11, bold=True, color='FFFFFF')
F_HDR   = Font(name=AR, size=10, bold=True)
F_B     = Font(name=AR, size=10)
F_BB    = Font(name=AR, size=10, bold=True)
F_NOTE  = Font(name=AR, size=9, italic=True, color='666666')
F_LINK  = Font(name=AR, size=9, color='0563C1', underline='single')
FI_GRP  = PatternFill('solid', fgColor='1F4E79')
FI_HDR  = PatternFill('solid', fgColor='DDEBF7')
FI_BLK  = PatternFill('solid', fgColor='FCE4D6')
FI_YEL  = PatternFill('solid', fgColor='FFF200')
_t = Side(style='thin', color='BFBFBF')
BD = Border(left=_t, right=_t, top=_t, bottom=_t)
CEN = Alignment(horizontal='center', vertical='center', wrap_text=True)


def hhmm(s):
    h, m = s.split(':')
    return datetime.time(int(h), int(m))


def build(cp, out):
    meta = cp['meta']
    SRC = meta['nguon']                       # thứ tự = thứ tự ưu tiên
    n = len(SRC)
    hdr = ["Cặp ngày", "Ngày bay", "Số hiệu chuyến", "Hãng", "Giờ khởi hành", "Giờ hạ cánh"]
    hdr += ["Giá — " + s for s in SRC]
    c_src0, c_src1 = 7, 6 + n                 # dải cột giá
    c_chot = c_src1 + 1
    hdr.append("Giá chốt")
    c_diff = 0
    if n >= 2:
        c_diff = c_chot + 1
        hdr.append("Chênh lệch max %")
    c_link = len(hdr) + 1; hdr.append("Link nguồn của Giá chốt")
    c_ts   = len(hdr) + 1; hdr.append("Thời điểm thu thập")
    NC = len(hdr)
    HDR_C = ["Loại phương án", "Chuyến đi (số hiệu + giờ)", "Giá đi",
             "Chuyến về (số hiệu + giờ)", "Giá về", "Tổng khứ hồi", "Ghi chú"]

    wb = Workbook(); ws = wb.active; ws.title = "Data"
    ws['A1'] = "%s — %s, %s, %s" % (meta['chang'], meta['khach'], meta['hang_ve'], meta['loai'])
    ws['A1'].font = F_TITLE
    ws['A2'] = ("Thời điểm thu thập: %s · Nguồn: %s · Khung giờ đi %s, về %s. "
                "Giá vé thay đổi liên tục — bảng này là ảnh chụp một thời điểm, không phải giá bảo đảm."
                % (meta['thu_thap'], ", ".join(SRC),
                   meta.get('khung_di', 'không lọc'), meta.get('khung_ve', 'không lọc')))
    ws['A2'].font = F_NOTE

    r, first_hdr, cf, summary = 4, None, [], []
    for P in cp['caps']:
        ws.cell(r, 1, "%s: %s → %s — %s" % (P['ten'], P['ngay_di'], P['ngay_ve'], P.get('thu', '')))
        for c in range(1, NC + 1):
            ws.cell(r, c).fill = FI_GRP; ws.cell(r, c).font = F_GRP
        r += 1
        blk = {}
        for leg, lbl, day in (('di', 'BLOCK A — Chuyến đi, khởi hành %s' % meta.get('khung_di', ''), P['ngay_di']),
                              ('ve', 'BLOCK B — Chuyến về, khởi hành %s' % meta.get('khung_ve', ''), P['ngay_ve'])):
            ws.cell(r, 1, lbl).font = F_HDR
            for c in range(1, NC + 1): ws.cell(r, c).fill = FI_BLK
            r += 1
            if first_hdr is None: first_hdr = r
            for i, h in enumerate(hdr, 1):
                c = ws.cell(r, i, h); c.font = F_HDR; c.fill = FI_HDR; c.border = BD; c.alignment = CEN
            r += 1
            start = r
            links = P[leg].get('link', {})
            for f in P[leg]['chuyen']:
                ws.cell(r, 1, P['ten']).font = F_B
                ws.cell(r, 2, day).font = F_B
                ws.cell(r, 3, f.get('sh') or '—').font = F_B
                ws.cell(r, 4, f['hang']).font = F_B
                for col, v in ((5, f['dep']), (6, f['arr'])):
                    cc = ws.cell(r, col, hhmm(v)); cc.number_format = 'hh:mm'
                    cc.font = F_B; cc.alignment = Alignment(horizontal='center')
                first_src = None
                for k, s in enumerate(SRC):
                    cc = ws.cell(r, 7 + k)
                    if s in f.get('gia', {}) and f['gia'][s] is not None:
                        cc.value = f['gia'][s]; cc.number_format = MONEY
                        if first_src is None: first_src = s
                    else:
                        cc.value = NA
                    cc.font = F_B
                expr = '"%s"' % NA
                for k in range(n - 1, -1, -1):
                    col = L(7 + k)
                    expr = 'IF(ISNUMBER({c}{r}),{c}{r},{e})'.format(c=col, r=r, e=expr)
                cc = ws.cell(r, c_chot, '=' + expr); cc.number_format = MONEY; cc.font = F_BB
                if c_diff:
                    cc = ws.cell(r, c_diff, '=IFERROR((MAX({a}{r}:{b}{r})-MIN({a}{r}:{b}{r}))/MIN({a}{r}:{b}{r}),"")'
                                 .format(a=L(c_src0), b=L(c_src1), r=r))
                    cc.number_format = '0.0%'; cc.font = F_B
                url = links.get(first_src or (SRC[0] if SRC else ''), '')
                lk = ws.cell(r, c_link, url or NA)
                if url: lk.hyperlink = url; lk.font = F_LINK
                else: lk.font = F_B
                ws.cell(r, c_ts, meta['thu_thap']).font = F_B
                for c in range(1, NC + 1): ws.cell(r, c).border = BD
                r += 1
            blk[leg] = (start, r - 1)
            cf.append((start, r - 1))
            r += 1

        # ---- Block C ----
        ws.cell(r, 1, 'BLOCK C — Phương án ghép').font = F_HDR
        for c in range(1, 8): ws.cell(r, c).fill = FI_BLK
        r += 1
        for i, h in enumerate(HDR_C, 1):
            c = ws.cell(r, i, h); c.font = F_HDR; c.fill = FI_HDR; c.border = BD; c.alignment = CEN
        r += 1

        def price(f):                                   # giá theo đúng thứ tự ưu tiên
            for s in SRC:
                v = f.get('gia', {}).get(s)
                if v is not None: return v
            return None

        def best(leg, hang=None):
            s0 = blk[leg][0]; b = None
            for i, f in enumerate(P[leg]['chuyen']):
                if hang and f['hang'] != hang: continue
                v = price(f)
                if v is None: continue
                if b is None or v < b[1]: b = (s0 + i, v, f)
            return b

        bA, bB = best('di'), best('ve')
        same = None
        for h in {f['hang'] for f in P['di']['chuyen']} & {f['hang'] for f in P['ve']['chuyen']}:
            a, b = best('di', h), best('ve', h)
            if a and b and (same is None or a[1] + b[1] < same[0]):
                same = (a[1] + b[1], h, a, b)
        opts = [("Rẻ nhất", bA, bB, "Chiều đi rẻ nhất + chiều về rẻ nhất (không cần cùng hãng)")]
        if same:
            opts.append(("Cùng hãng rẻ nhất", same[2], same[3],
                         "Hai chặng cùng %s — tiện đổi/hoãn vé" % same[1]))
        for name, A, B, note in opts:
            if not (A and B): continue
            ws.cell(r, 1, name).font = F_BB
            for col, X in ((2, A), (4, B)):
                ws.cell(r, col, '=C{0}&" "&TEXT(E{0},"hh:mm")'.format(X[0])).font = F_B
            for col, X in ((3, A), (5, B)):
                cc = ws.cell(r, col, '=%s%d' % (L(c_chot), X[0])); cc.number_format = MONEY; cc.font = F_B
            cc = ws.cell(r, 6, '=C{0}+E{0}'.format(r)); cc.number_format = MONEY; cc.font = F_BB
            ws.cell(r, 7, note).font = F_B
            for c in range(1, 8): ws.cell(r, c).border = BD
            if name == "Rẻ nhất":
                summary.append(dict(P=P, row=r, A=A, B=B))
            r += 1
        ws.cell(r, 1, "Vé nội địa VN không có discount khứ hồi — tổng = đi + về.").font = F_NOTE
        r += 2

    for s, e in cf:
        ws.conditional_formatting.add(
            "{c}{s}:{c}{e}".format(c=L(c_chot), s=s, e=e),
            FormulaRule(formula=['AND(ISNUMBER(${c}{s}),${c}{s}=MIN(${c}${s}:${c}${e}))'
                                 .format(c=L(c_chot), s=s, e=e)], fill=FI_YEL, stopIfTrue=False))
    ws.conditional_formatting.add("A4:G%d" % r,
        FormulaRule(formula=['$A4="Rẻ nhất"'], fill=FI_YEL, stopIfTrue=False))
    ws.freeze_panes = "A4"
    if first_hdr: ws.auto_filter.ref = "A%d:%s%d" % (first_hdr, L(NC), r)
    for i, w in enumerate([10, 12, 14, 20, 13, 13] + [16] * n + [15] + ([14] if c_diff else []) + [42, 18], 1):
        ws.column_dimensions[L(i)].width = w

    # ---------------- Summary ----------------
    s2 = wb.create_sheet("Summary")
    for i, h in enumerate(["Cặp ngày", "Thứ (đi → về)", "Tổng vé rẻ nhất", "Hãng chiều đi", "Giờ đi",
                           "Hãng chiều về", "Giờ về", "Nguồn", "Link nguồn"], 1):
        c = s2.cell(1, i, h); c.font = F_HDR; c.fill = FI_HDR; c.border = BD; c.alignment = CEN
    rr = 2
    for it in sorted(summary, key=lambda x: x['A'][1] + x['B'][1]):
        P, A, B = it['P'], it['A'], it['B']
        src = next((s for s in SRC if A[2].get('gia', {}).get(s) is not None), SRC[0])
        s2.cell(rr, 1, "%s (%s → %s)" % (P['ten'], P['ngay_di'], P['ngay_ve'])).font = F_B
        s2.cell(rr, 2, P.get('thu', '')).font = F_B
        cc = s2.cell(rr, 3, "=Data!F%d" % it['row']); cc.number_format = MONEY; cc.font = F_BB
        s2.cell(rr, 4, "%s (%s)" % (A[2]['hang'], A[2].get('sh') or '—')).font = F_B
        cc = s2.cell(rr, 5, hhmm(A[2]['dep'])); cc.number_format = 'hh:mm'; cc.font = F_B
        s2.cell(rr, 6, "%s (%s)" % (B[2]['hang'], B[2].get('sh') or '—')).font = F_B
        cc = s2.cell(rr, 7, hhmm(B[2]['dep'])); cc.number_format = 'hh:mm'; cc.font = F_B
        s2.cell(rr, 8, src).font = F_B
        u = P['di'].get('link', {}).get(src, '')
        lk = s2.cell(rr, 9, u or NA)
        if u: lk.hyperlink = u; lk.font = F_LINK
        else: lk.font = F_B
        for c in range(1, 10): s2.cell(rr, c).border = BD
        rr += 1
    if rr > 2:
        s2.conditional_formatting.add("A2:I%d" % (rr - 1),
            FormulaRule(formula=['$C2=MIN($C$2:$C$%d)' % (rr - 1)], fill=FI_YEL, stopIfTrue=False))
        s2.auto_filter.ref = "A1:I%d" % (rr - 1)
    s2.freeze_panes = "A2"
    for i, w in enumerate([26, 16, 18, 24, 10, 24, 10, 16, 42], 1): s2.column_dimensions[L(i)].width = w

    # ---------------- Log ----------------
    s3 = wb.create_sheet("Log")
    for i, h in enumerate(["Thời điểm", "Cặp ngày", "Nguồn", "Trạng thái", "Lý do chi tiết", "Ghi chú"], 1):
        c = s3.cell(1, i, h); c.font = F_HDR; c.fill = FI_HDR; c.border = BD; c.alignment = CEN
    rr = 2
    for e in cp.get('log', []):
        for i, k in enumerate(['thoi_diem', 'cap', 'nguon', 'trang_thai', 'ly_do', 'ghi_chu'], 1):
            c = s3.cell(rr, i, e.get(k, '')); c.font = F_B; c.border = BD
            c.alignment = Alignment(wrap_text=True, vertical='top')
        rr += 1
    s3.freeze_panes = "A2"
    if rr > 2: s3.auto_filter.ref = "A1:F%d" % (rr - 1)
    for i, w in enumerate([22, 16, 30, 20, 80, 42], 1): s3.column_dimensions[L(i)].width = w

    wb.save(out)
    nf = sum(len(P[l]['chuyen']) for P in cp['caps'] for l in ('di', 've'))
    have = sum(1 for P in cp['caps'] for l in ('di', 've') for f in P[l]['chuyen']
               for s in SRC if f.get('gia', {}).get(s) is not None)
    print(json.dumps({"file": out, "chuyen": nf, "o_co_gia": have,
                      "o_tong": nf * len(SRC),
                      "phu_%": round(100.0 * have / max(1, nf * len(SRC)), 1)},
                     ensure_ascii=False))


if __name__ == '__main__':
    if len(sys.argv) != 3:
        sys.exit(__doc__)
    build(json.load(open(sys.argv[1], encoding='utf-8')), sys.argv[2])
